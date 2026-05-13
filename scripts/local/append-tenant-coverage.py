#!/usr/bin/env python3
# =====================================================================
# append-tenant-coverage.py
# 按 "租户分工" 策略向 ta / tb / tc-tenant-config-package-test.xlsx 追加
# v4 场景覆盖，补齐枚举 + 场景长尾：
#   - ta (零售)：IMPORT FEEDBACK 全链（6 stage） + EXPORT 全链（5 stage） + LOCAL channel
#   - tb (金融)：DISPATCH 全链（6 stage） + API channel（区别于已有 API_PUSH）
#   - tc (风控)：GATEWAY ALL / N_OF join 两个 workflow + FAILURE / CONDITION 边
# default-tenant 继续承载 MIXED + 探针，由 gen-default-tenant-excel.py 管。
#
# 幂等：追加前按 "主键" 查重跳过；重复跑不会产生重复行。
# 用法：python3 scripts/local/append-tenant-coverage.py
# =====================================================================
import json
from openpyxl import load_workbook

BASE = "docs/test-data/test-full-coverage-import-suite/"

def row_exists(ws, key_col_idx, key_value):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[key_col_idx] == key_value:
            return True
    return False

def composite_row_exists(ws, indices, values):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(row[i] == v for i, v in zip(indices, values)):
            return True
    return False

def append_if_new(ws, key_cols, keys, row_values):
    """key_cols: list of column indices identifying the row; keys: values at those indices."""
    if composite_row_exists(ws, key_cols, keys):
        return False
    ws.append(row_values)
    return True

# ──────────────────────────────────────────────────────────────
# ta (零售)
# ──────────────────────────────────────────────────────────────
def update_ta():
    path = BASE + "ta-tenant-config-package-test.xlsx"
    wb = load_workbook(path)

    # 新 IMPORT job：TA_IMPORT_ORDER（ORDER 业务，独立 6-stage 链路，避免改动既有 TA_IMPORT_CUSTOMER 的 step_order）
    # job_definition cols: tenant_id, job_code, job_name, job_type, biz_type, queue_code, worker_group,
    #   schedule_type, schedule_expr, calendar_code, window_code, retry_policy, retry_max_count,
    #   timeout_seconds, shard_strategy, execution_handler, param_schema, default_params, enabled, description
    jd = wb["job_definition"]
    append_if_new(jd, [1], ["TA_IMPORT_ORDER"], [
        "ta","TA_IMPORT_ORDER","零售订单导入","IMPORT","ORDER","ta-import-queue","import",
        "CRON","0 3 * * *","default-calendar","always-open","EXPONENTIAL",3,3600,"STATIC",
        "orderImportHandler","{}","{}","TRUE","FEEDBACK 全链覆盖样本",
    ])
    # EXPORT 已有 TA_EXPORT_REPORT 但无 pipeline；不改 job，只为其补 pipeline_definition + 5 stage

    # file_channel_config cols: tenant_id, channel_code, channel_name, channel_type, target_endpoint,
    #   auth_type, config_json, receipt_policy, timeout_seconds, enabled
    ch = wb["file_channel_config"]
    append_if_new(ch, [1], ["ta_local_archive"], [
        "ta","ta_local_archive","本地归档落盘","LOCAL","/tmp/batch/ta-archive","NONE",
        json.dumps({"target_endpoint":"/tmp/batch/ta-archive","local_directory":"/tmp/batch/ta-archive",
                    "local_file_name":"customer-{bizDate}.csv"}, ensure_ascii=False),
        "NONE",30,"TRUE",
    ])

    # pipeline_definition cols: tenant_id, job_code, pipeline_name, pipeline_type, biz_type,
    #   worker_group, version, enabled, description
    pd = wb["pipeline_definition"]
    append_if_new(pd, [1], ["TA_IMPORT_ORDER"], [
        "ta","TA_IMPORT_ORDER","订单 6-stage 导入流水线","IMPORT","ORDER","import",
        1,"TRUE","RECEIVE→PREPROCESS→PARSE→VALIDATE→LOAD→FEEDBACK 全链",
    ])
    append_if_new(pd, [1], ["TA_EXPORT_REPORT"], [
        "ta","TA_EXPORT_REPORT","报表 5-stage 导出流水线","EXPORT","REPORT","export",
        1,"TRUE","PREPARE→GENERATE→STORE→REGISTER→COMPLETE 全链",
    ])

    # pipeline_step_definition cols: job_code, version, step_code, step_name, stage_code, step_order,
    #   impl_code, step_params, timeout_seconds, retry_policy, retry_max_count, enabled
    ps = wb["pipeline_step_definition"]
    # impl_code 必须匹配 worker 启动时上报的 step_registry：模式 = "{MODULE}_{STAGE}"
    # 例：IMPORT/RECEIVE → IMPORT_RECEIVE；validator 在 preview 时校验
    import_steps = [
        ("STEP_RECEIVE",   "接收文件","RECEIVE",   1,"IMPORT_RECEIVE"),
        ("STEP_PREPROCESS","预处理",  "PREPROCESS",2,"IMPORT_PREPROCESS"),
        ("STEP_PARSE",     "解析",    "PARSE",     3,"IMPORT_PARSE"),
        ("STEP_VALIDATE",  "校验",    "VALIDATE",  4,"IMPORT_VALIDATE"),
        ("STEP_LOAD",      "入库",    "LOAD",      5,"IMPORT_LOAD"),
        ("STEP_FEEDBACK",  "回执写回","FEEDBACK",  6,"IMPORT_FEEDBACK"),
    ]
    for sc, sn, stage, order, impl in import_steps:
        append_if_new(ps, [0,2], ["TA_IMPORT_ORDER", sc], [
            "TA_IMPORT_ORDER",1,sc,sn,stage,order,impl,"{}",300,"FIXED",2,"TRUE",
        ])
    export_steps = [
        ("STEP_PREPARE",  "生成前准备","PREPARE",  1,"EXPORT_PREPARE"),
        ("STEP_GENERATE", "生成文件",  "GENERATE", 2,"EXPORT_GENERATE"),
        ("STEP_STORE",    "落盘存档",  "STORE",    3,"EXPORT_STORE"),
        ("STEP_REGISTER", "登记 FileRecord","REGISTER",4,"EXPORT_REGISTER"),
        ("STEP_COMPLETE", "完结",      "COMPLETE", 5,"EXPORT_COMPLETE"),
    ]
    for sc, sn, stage, order, impl in export_steps:
        append_if_new(ps, [0,2], ["TA_EXPORT_REPORT", sc], [
            "TA_EXPORT_REPORT",1,sc,sn,stage,order,impl,"{}",300,"FIXED",2,"TRUE",
        ])

    wb.save(path)
    print(f"ta: updated {path}")

# ──────────────────────────────────────────────────────────────
# tb (金融)
# ──────────────────────────────────────────────────────────────
def update_tb():
    path = BASE + "tb-tenant-config-package-test.xlsx"
    wb = load_workbook(path)

    jd = wb["job_definition"]
    append_if_new(jd, [1], ["TB_DISPATCH_SETTLE"], [
        "tb","TB_DISPATCH_SETTLE","对账单派发","DISPATCH","STATEMENT","tb-dispatch-queue","dispatch",
        "CRON","0 0 8 * * ?","default-calendar","always-open","EXPONENTIAL",5,3600,"NONE",
        "statementDispatchHandler","{}","{}","TRUE","DISPATCH 全链 6-stage 覆盖样本",
    ])

    ch = wb["file_channel_config"]
    append_if_new(ch, [1], ["tb_api_ingest"], [
        "tb","tb_api_ingest","核心系统 API 同步入站","API","http://mockserver:1080/tb/ingest",
        "TOKEN",
        json.dumps({"url":"http://mockserver:1080/tb/ingest","method":"POST",
                    "tokenHeader":"X-API-Token","target_endpoint":"http://mockserver:1080/tb/ingest"},
                   ensure_ascii=False),
        "SYNC",15,"TRUE",
    ])

    pd = wb["pipeline_definition"]
    append_if_new(pd, [1], ["TB_DISPATCH_SETTLE"], [
        "tb","TB_DISPATCH_SETTLE","派发全链流水线","DISPATCH","STATEMENT","dispatch",
        1,"TRUE","PREPARE→DISPATCH→ACK→RETRY→COMPENSATE→COMPLETE 全链",
    ])

    ps = wb["pipeline_step_definition"]
    dispatch_steps = [
        ("STEP_PREPARE",   "分发前准备","PREPARE",   1,"DISPATCH_PREPARE"),
        ("STEP_DISPATCH",  "实际下发",  "DISPATCH",  2,"DISPATCH_DISPATCH"),
        ("STEP_ACK",       "回执确认",  "ACK",       3,"DISPATCH_ACK"),
        ("STEP_RETRY",     "失败重试",  "RETRY",     4,"DISPATCH_RETRY"),
        ("STEP_COMPENSATE","补偿冲正",  "COMPENSATE",5,"DISPATCH_COMPENSATE"),
        ("STEP_COMPLETE",  "完结",      "COMPLETE",  6,"DISPATCH_COMPLETE"),
    ]
    for sc, sn, stage, order, impl in dispatch_steps:
        append_if_new(ps, [0,2], ["TB_DISPATCH_SETTLE", sc], [
            "TB_DISPATCH_SETTLE",1,sc,sn,stage,order,impl,"{}",300,"FIXED",2,"TRUE",
        ])

    wb.save(path)
    print(f"tb: updated {path}")

# ──────────────────────────────────────────────────────────────
# tc (风控)
# ──────────────────────────────────────────────────────────────
def update_tc():
    path = BASE + "tc-tenant-config-package-test.xlsx"
    wb = load_workbook(path)

    jd = wb["job_definition"]
    append_if_new(jd, [1], ["TC_WF_GATEWAY_ALL"], [
        "tc","TC_WF_GATEWAY_ALL","GATEWAY ALL join 工作流","WORKFLOW","RISK",None,None,
        "MANUAL",None,"default-calendar","always-open","NONE",0,14400,"NONE",
        None,"{}","{}","TRUE","GATEWAY joinMode=ALL 覆盖",
    ])
    append_if_new(jd, [1], ["TC_WF_GATEWAY_N_OF"], [
        "tc","TC_WF_GATEWAY_N_OF","GATEWAY N_OF join 工作流","WORKFLOW","RISK",None,None,
        "MANUAL",None,"default-calendar","always-open","NONE",0,14400,"NONE",
        None,"{}","{}","TRUE","GATEWAY joinMode=N_OF + joinThreshold=2 覆盖",
    ])

    wf = wb["workflow_definition"]
    append_if_new(wf, [1], ["TC_WF_GATEWAY_ALL"], [
        "tc","TC_WF_GATEWAY_ALL","风控 GATEWAY ALL 工作流","DAG",1,"TRUE","FORK→3 branch→MERGE(ALL)→END",
    ])
    append_if_new(wf, [1], ["TC_WF_GATEWAY_N_OF"], [
        "tc","TC_WF_GATEWAY_N_OF","风控 GATEWAY N_OF 工作流","DAG",1,"TRUE","FORK→3 branch→MERGE(N_OF=2)→END + FAILURE / CONDITION 边",
    ])

    wn = wb["workflow_node"]
    # node_params JSON strings
    def np(d): return json.dumps(d, ensure_ascii=False)
    # workflow_node cols: tenant_id, workflow_code, workflow_version, node_code, node_name, node_type,
    #   related_job_code, related_pipeline_code, worker_group, window_code, node_order, retry_policy,
    #   retry_max_count, timeout_seconds, node_params, enabled
    all_nodes = [
        ("TC_WF_GATEWAY_ALL","START",    "Start",    "START",  None,                 0, {"entry":True}),
        ("TC_WF_GATEWAY_ALL","FORK",     "Fork",     "GATEWAY",None,                 1, {}),
        ("TC_WF_GATEWAY_ALL","BRANCH_A", "Branch A", "JOB",    "TC_IMPORT_RISK_SCORE",2, {"branch":"A"}),
        ("TC_WF_GATEWAY_ALL","BRANCH_B", "Branch B", "JOB",    "TC_EXPORT_RISK_ALERT",3, {"branch":"B"}),
        ("TC_WF_GATEWAY_ALL","BRANCH_C", "Branch C", "JOB",    "TC_DISPATCH_REVIEW",  4, {"branch":"C"}),
        ("TC_WF_GATEWAY_ALL","MERGE",    "Merge ALL","GATEWAY",None,                 5, {"joinMode":"ALL"}),
        ("TC_WF_GATEWAY_ALL","END",      "End",      "END",    None,                 6, {"entry":False}),

        ("TC_WF_GATEWAY_N_OF","START",    "Start",     "START",  None,                 0, {"entry":True}),
        ("TC_WF_GATEWAY_N_OF","FORK",     "Fork",      "GATEWAY",None,                 1, {}),
        ("TC_WF_GATEWAY_N_OF","BRANCH_A", "Branch A",  "JOB",    "TC_IMPORT_RISK_SCORE",2, {"branch":"A"}),
        ("TC_WF_GATEWAY_N_OF","BRANCH_B", "Branch B",  "JOB",    "TC_EXPORT_RISK_ALERT",3, {"branch":"B"}),
        ("TC_WF_GATEWAY_N_OF","BRANCH_C", "Branch C",  "JOB",    "TC_DISPATCH_REVIEW",  4, {"branch":"C"}),
        ("TC_WF_GATEWAY_N_OF","MERGE",    "Merge N_OF","GATEWAY",None,                 5, {"joinMode":"N_OF","joinThreshold":2}),
        ("TC_WF_GATEWAY_N_OF","END",      "End",       "END",    None,                 6, {"entry":False}),
        ("TC_WF_GATEWAY_N_OF","FALLBACK", "Fallback",  "JOB",    "TC_DISPATCH_REVIEW",  7, {"branch":"fallback"}),
    ]
    for wc, nc, nn, nt, rjc, order, params in all_nodes:
        append_if_new(wn, [1,3], [wc, nc], [
            "tc",wc,1,nc,nn,nt,rjc,None,None,None,order,"NONE",0,0,np(params),"TRUE",
        ])

    we = wb["workflow_edge"]
    # workflow_edge cols: tenant_id, workflow_code, workflow_version, from_node_code, to_node_code,
    #   edge_type, condition_expr, enabled
    edges = [
        ("TC_WF_GATEWAY_ALL","START","FORK","ALWAYS",None),
        ("TC_WF_GATEWAY_ALL","FORK","BRANCH_A","ALWAYS",None),
        ("TC_WF_GATEWAY_ALL","FORK","BRANCH_B","ALWAYS",None),
        ("TC_WF_GATEWAY_ALL","FORK","BRANCH_C","ALWAYS",None),
        ("TC_WF_GATEWAY_ALL","BRANCH_A","MERGE","SUCCESS",None),
        ("TC_WF_GATEWAY_ALL","BRANCH_B","MERGE","SUCCESS",None),
        ("TC_WF_GATEWAY_ALL","BRANCH_C","MERGE","SUCCESS",None),
        ("TC_WF_GATEWAY_ALL","MERGE","END","ALWAYS",None),
        # N_OF 工作流：补 FAILURE / CONDITION 边类型
        ("TC_WF_GATEWAY_N_OF","START","FORK","ALWAYS",None),
        ("TC_WF_GATEWAY_N_OF","FORK","BRANCH_A","ALWAYS",None),
        ("TC_WF_GATEWAY_N_OF","FORK","BRANCH_B","CONDITION","${bizDate != null}"),
        ("TC_WF_GATEWAY_N_OF","FORK","BRANCH_C","ALWAYS",None),
        ("TC_WF_GATEWAY_N_OF","BRANCH_A","MERGE","SUCCESS",None),
        ("TC_WF_GATEWAY_N_OF","BRANCH_B","MERGE","SUCCESS",None),
        ("TC_WF_GATEWAY_N_OF","BRANCH_C","MERGE","SUCCESS",None),
        ("TC_WF_GATEWAY_N_OF","BRANCH_C","FALLBACK","FAILURE",None),
        ("TC_WF_GATEWAY_N_OF","FALLBACK","END","ALWAYS",None),
        ("TC_WF_GATEWAY_N_OF","MERGE","END","ALWAYS",None),
    ]
    for wc, fnc, tnc, et, cond in edges:
        append_if_new(we, [1,3,4], [wc, fnc, tnc], [
            "tc",wc,1,fnc,tnc,et,cond,"TRUE",
        ])

    wb.save(path)
    print(f"tc: updated {path}")

if __name__ == "__main__":
    update_ta()
    update_tb()
    update_tc()
