#!/usr/bin/env python3
"""Idempotent patches for sim-e2e fixture xlsx workbooks.

Patches applied to ta/tb/tc-tenant-config-package-test.xlsx:
  1. job_definition.schedule_expr Linux 5-field -> Quartz 6-field
     e.g.  '0 2 * * *'  ->  '0 0 2 * * ?'
  2. file_channel_config.config_json SFTP key rename
     host/port/username/password/remotePath -> sftp_*  (also legacy snake forms)
  3. workflow_node: ensure START / END boundary nodes exist
     - rename node_code NODE_START/NODE_END -> START/END if present
     - if missing, append START (node_type=START) + END (node_type=END)
     - workflow_edge wired START -> first JOB node, last JOB node -> END
  4. job_definition.execution_mode required by current Console package validator.
     Add column when absent and fill FULL for every job row.
  5. file_template_config JSON/SQL preview hardening:
     - field_mappings entries must carry name.
     - export default_query_sql must list columns explicitly.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SUITE_DIR = ROOT / "docs/test-data/test-full-coverage-import-suite"
TARGETS = ["ta-tenant-config-package-test.xlsx",
           "tb-tenant-config-package-test.xlsx",
           "tc-tenant-config-package-test.xlsx",
           "default-tenant-config-package-test.xlsx"]

LINUX_5FIELD_RE = re.compile(r"^\s*(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s*$")

SFTP_KEY_MAP = {
    "host": "sftp_host",
    "sftpHost": "sftp_host",
    "port": "sftp_port",
    "sftpPort": "sftp_port",
    "username": "sftp_username",
    "sftpUsername": "sftp_username",
    "password": "sftp_password",
    "sftpPassword": "sftp_password",
    "remotePath": "sftp_remote_path",
    "sftpRemotePath": "sftp_remote_path",
    "remote_path": "sftp_remote_path",
}

TEMPLATE_MAPPINGS: dict[str, list[dict[str, object]]] = {
    "TA_IMPORT_CUSTOMER_TPL": [
        {"name": "customer_no", "targetColumn": "customer_no", "type": "STRING", "required": True},
        {"name": "customer_name", "targetColumn": "customer_name", "type": "STRING", "required": True},
        {"name": "customer_type", "targetColumn": "customer_type", "type": "STRING", "required": True},
        {"name": "certificate_no", "targetColumn": "certificate_no", "type": "STRING"},
        {"name": "mobile_no", "targetColumn": "mobile_no", "type": "STRING"},
        {"name": "email", "targetColumn": "email", "type": "STRING"},
        {"name": "status", "targetColumn": "status", "type": "STRING", "required": True},
    ],
    "TA_IMPORT_ORDER_TPL": [
        {"name": "customer_no", "targetColumn": "customer_no", "type": "STRING", "required": True},
        {"name": "customer_name", "targetColumn": "customer_name", "type": "STRING", "required": True},
        {"name": "customer_type", "targetColumn": "customer_type", "type": "STRING", "required": True},
        {"name": "certificate_no", "targetColumn": "certificate_no", "type": "STRING"},
        {"name": "mobile_no", "targetColumn": "mobile_no", "type": "STRING"},
        {"name": "email", "targetColumn": "email", "type": "STRING"},
        {"name": "status", "targetColumn": "status", "type": "STRING", "required": True},
    ],
    "TB_IMPORT_TRANSACTION_TPL": [
        {"name": "txn_no", "targetColumn": "txn_no", "type": "STRING", "required": True},
        {"name": "account_no", "targetColumn": "account_no", "type": "STRING", "required": True},
        {"name": "txn_type", "targetColumn": "txn_type", "type": "STRING", "required": True},
        {"name": "amount", "targetColumn": "amount", "type": "DECIMAL", "required": True},
        {"name": "currency_code", "targetColumn": "currency_code", "type": "STRING", "required": True},
        {"name": "txn_date", "targetColumn": "txn_date", "type": "DATE", "required": True, "format": "yyyy-MM-dd"},
        {"name": "remark", "targetColumn": "remark", "type": "STRING"},
    ],
    "TC_IMPORT_RISK_SCORE_TPL": [
        {"name": "entity_id", "targetColumn": "entity_id", "type": "STRING", "required": True},
        {"name": "entity_type", "targetColumn": "entity_type", "type": "STRING", "required": True},
        {"name": "score_value", "targetColumn": "score_value", "type": "INTEGER", "required": True},
        {"name": "score_band", "targetColumn": "score_band", "type": "STRING", "required": True},
        {"name": "score_date", "targetColumn": "score_date", "type": "DATE", "required": True, "format": "yyyy-MM-dd"},
    ],
    "TA_EXPORT_REPORT_TPL": [
        {"name": "id", "sourceColumn": "id", "header": "ID"},
        {"name": "tenant_id", "sourceColumn": "tenant_id", "header": "Tenant"},
        {"name": "customer_no", "sourceColumn": "customer_no", "header": "Customer No"},
        {"name": "customer_name", "sourceColumn": "customer_name", "header": "Customer Name"},
        {"name": "customer_type", "sourceColumn": "customer_type", "header": "Customer Type"},
        {"name": "certificate_no", "sourceColumn": "certificate_no", "header": "Certificate No"},
        {"name": "mobile_no", "sourceColumn": "mobile_no", "header": "Mobile"},
        {"name": "email", "sourceColumn": "email", "header": "Email"},
        {"name": "status", "sourceColumn": "status", "header": "Status"},
    ],
    "TB_EXPORT_STATEMENT_TPL": [
        {"name": "id", "sourceColumn": "id", "header": "ID"},
        {"name": "tenant_id", "sourceColumn": "tenant_id", "header": "Tenant"},
        {"name": "txn_no", "sourceColumn": "txn_no", "header": "Txn No"},
        {"name": "account_no", "sourceColumn": "account_no", "header": "Account No"},
        {"name": "txn_type", "sourceColumn": "txn_type", "header": "Txn Type"},
        {"name": "amount", "sourceColumn": "amount", "header": "Amount", "type": "DECIMAL"},
        {"name": "currency_code", "sourceColumn": "currency_code", "header": "Currency"},
        {"name": "txn_date", "sourceColumn": "txn_date", "header": "Txn Date", "type": "DATE", "format": "yyyy-MM-dd"},
        {"name": "remark", "sourceColumn": "remark", "header": "Remark"},
    ],
    "TC_EXPORT_RISK_ALERT_TPL": [
        {"name": "id", "sourceColumn": "id", "header": "ID"},
        {"name": "tenant_id", "sourceColumn": "tenant_id", "header": "Tenant"},
        {"name": "entity_id", "sourceColumn": "entity_id", "header": "Entity ID"},
        {"name": "entity_type", "sourceColumn": "entity_type", "header": "Entity Type"},
        {"name": "score_value", "sourceColumn": "score_value", "header": "Score"},
        {"name": "score_band", "sourceColumn": "score_band", "header": "Band"},
        {"name": "score_date", "sourceColumn": "score_date", "header": "Score Date", "type": "DATE", "format": "yyyy-MM-dd"},
    ],
}

EXPORT_SQL: dict[str, str] = {
    "TA_EXPORT_REPORT_TPL":
        "SELECT id, tenant_id, customer_no, customer_name, customer_type, certificate_no, mobile_no, email, status "
        "FROM biz.customer_account WHERE tenant_id = :tenantId AND (:batchNo IS NULL OR :batchNo IS NOT NULL)",
    "TB_EXPORT_STATEMENT_TPL":
        "SELECT id, tenant_id, txn_no, account_no, txn_type, amount, currency_code, txn_date, remark "
        "FROM biz.transaction WHERE tenant_id = :tenantId AND (:batchNo IS NULL OR :batchNo IS NOT NULL)",
    "TC_EXPORT_RISK_ALERT_TPL":
        "SELECT id, tenant_id, entity_id, entity_type, score_value, score_band, score_date "
        "FROM biz.risk_score WHERE tenant_id = :tenantId AND (:batchNo IS NULL OR :batchNo IS NOT NULL)",
}


def linux5_to_quartz6(expr: str) -> str | None:
    if not expr:
        return None
    s = str(expr).strip()
    if not s:
        return None
    parts = s.split()
    if len(parts) in (6, 7):
        return None  # already quartz
    m = LINUX_5FIELD_RE.match(s)
    if not m:
        return None
    minute, hour, dom, month, dow = m.groups()
    # Quartz: sec min hour day-of-month month day-of-week [year]
    # Linux DOW 0-6 (Sun=0) ; Quartz DOW 1-7 (Sun=1). If '*' or '?', keep '?'.
    if dom == "*" and dow == "*":
        q_dom, q_dow = "*", "?"
    elif dow == "*":
        q_dom, q_dow = dom, "?"
    elif dom == "*":
        q_dom, q_dow = "?", dow
    else:
        q_dom, q_dow = dom, "?"
    return f"0 {minute} {hour} {q_dom} {month} {q_dow}"


def patch_schedule_expr(ws) -> int:
    headers = [c.value for c in ws[1]]
    if "schedule_expr" not in headers:
        return 0
    col = headers.index("schedule_expr") + 1
    type_col = headers.index("schedule_type") + 1 if "schedule_type" in headers else None
    fixed = 0
    for row in ws.iter_rows(min_row=2):
        if type_col is not None and (row[type_col - 1].value or "").strip().upper() != "CRON":
            continue
        cell = row[col - 1]
        new = linux5_to_quartz6(cell.value)
        if new is not None:
            cell.value = new
            fixed += 1
    return fixed


def ensure_execution_mode(ws) -> int:
    headers = [c.value for c in ws[1]]
    if "execution_mode" in headers:
        col = headers.index("execution_mode") + 1
    else:
        col = len(headers) + 1
        ws.cell(row=1, column=col).value = "execution_mode"
    fixed = 0
    for row in range(2, ws.max_row + 1):
        tenant = ws.cell(row=row, column=1).value
        job_code = ws.cell(row=row, column=2).value
        if tenant in (None, "") and job_code in (None, ""):
            continue
        cell = ws.cell(row=row, column=col)
        if cell.value in (None, ""):
            cell.value = "FULL"
            fixed += 1
    return fixed


def patch_channel_config_json(ws) -> int:
    headers = [c.value for c in ws[1]]
    if "config_json" not in headers or "channel_type" not in headers:
        return 0
    cj_col = headers.index("config_json") + 1
    ct_col = headers.index("channel_type") + 1
    fixed = 0
    for row in ws.iter_rows(min_row=2):
        if (row[ct_col - 1].value or "").strip().upper() != "SFTP":
            continue
        raw = row[cj_col - 1].value
        if not raw:
            continue
        try:
            obj = json.loads(raw)
        except Exception:
            continue
        if not isinstance(obj, dict):
            continue
        changed = False
        new_obj = {}
        for k, v in obj.items():
            nk = SFTP_KEY_MAP.get(k, k)
            if nk != k:
                changed = True
            new_obj[nk] = v
        if changed:
            row[cj_col - 1].value = json.dumps(new_obj, ensure_ascii=False)
            fixed += 1
    return fixed


def patch_file_template_config(ws) -> int:
    headers = [c.value for c in ws[1]]
    required = {"template_code", "template_type", "field_mappings", "default_query_sql"}
    if not required.issubset(set(headers)):
        return 0
    idx = {h: i for i, h in enumerate(headers)}
    fixed = 0
    for row in ws.iter_rows(min_row=2):
        template_code = row[idx["template_code"]].value
        template_type = (row[idx["template_type"]].value or "").strip().upper()
        if not template_code:
            continue

        mappings = TEMPLATE_MAPPINGS.get(template_code)
        if mappings is not None:
            rendered = json.dumps(mappings, ensure_ascii=False, separators=(",", ":"))
            cell = row[idx["field_mappings"]]
            if cell.value != rendered:
                cell.value = rendered
                fixed += 1
        else:
            cell = row[idx["field_mappings"]]
            patched = ensure_mapping_names(cell.value, template_type)
            if patched is not None and patched != cell.value:
                cell.value = patched
                fixed += 1

        sql = EXPORT_SQL.get(template_code)
        if sql is not None:
            cell = row[idx["default_query_sql"]]
            if cell.value != sql:
                cell.value = sql
                fixed += 1
    return fixed


def ensure_mapping_names(raw: object, template_type: str) -> str | None:
    if raw in (None, ""):
        return None
    try:
        mappings = json.loads(str(raw))
    except Exception:
        return None
    if not isinstance(mappings, list):
        return None
    changed = False
    for mapping in mappings:
        if not isinstance(mapping, dict):
            continue
        name = str(mapping.get("name") or "").strip()
        if name:
            continue
        candidate = (
            mapping.get("source")
            or mapping.get("targetColumn")
            or mapping.get("sourceColumn")
            or mapping.get("target")
        )
        if candidate in (None, ""):
            continue
        mapping["name"] = str(candidate)
        if template_type == "EXPORT" and "sourceColumn" not in mapping:
            mapping["sourceColumn"] = str(candidate)
        changed = True
    return json.dumps(mappings, ensure_ascii=False, separators=(",", ":")) if changed else None


def ensure_workflow_boundary_nodes(wb) -> tuple[int, int]:
    """Make sure each (workflow_code, workflow_version) has node_code=START+END.

    Renames NODE_START/NODE_END -> START/END; appends missing; wires edges.
    """
    if "workflow_node" not in wb.sheetnames:
        return 0, 0
    node_ws = wb["workflow_node"]
    edge_ws = wb["workflow_edge"] if "workflow_edge" in wb.sheetnames else None
    n_headers = [c.value for c in node_ws[1]]
    nidx = {h: i for i, h in enumerate(n_headers)}

    # group rows by (tenant, code, version)
    groups: dict[tuple, list] = {}
    for row in node_ws.iter_rows(min_row=2):
        if row[nidx["tenant_id"]].value in (None, ""):
            continue
        key = (row[nidx["tenant_id"]].value,
               row[nidx["workflow_code"]].value,
               row[nidx["workflow_version"]].value)
        groups.setdefault(key, []).append(row)

    renamed = 0
    appended = 0
    edges_added = 0

    for key, rows in groups.items():
        tenant, wf_code, wf_ver = key
        has_start = False
        has_end = False
        first_job = None
        last_job = None
        for r in rows:
            nc = r[nidx["node_code"]].value
            nt = (r[nidx["node_type"]].value or "").upper()
            if nc == "NODE_START":
                r[nidx["node_code"]].value = "START"
                r[nidx["node_type"]].value = "START"
                nc, nt = "START", "START"
                renamed += 1
            if nc == "NODE_END":
                r[nidx["node_code"]].value = "END"
                r[nidx["node_type"]].value = "END"
                nc, nt = "END", "END"
                renamed += 1
            if nc == "START" or nt == "START":
                has_start = True
            if nc == "END" or nt == "END":
                has_end = True
            if nt == "JOB":
                if first_job is None:
                    first_job = nc
                last_job = nc

        max_order = 0
        for r in rows:
            o = r[nidx["node_order"]].value
            if isinstance(o, int) and o > max_order:
                max_order = o

        appended_codes: list[str] = []
        if not has_start:
            new_row = [None] * len(n_headers)
            new_row[nidx["tenant_id"]] = tenant
            new_row[nidx["workflow_code"]] = wf_code
            new_row[nidx["workflow_version"]] = wf_ver
            new_row[nidx["node_code"]] = "START"
            new_row[nidx["node_name"]] = "开始"
            new_row[nidx["node_type"]] = "START"
            new_row[nidx["node_order"]] = 0
            if "retry_policy" in nidx:
                new_row[nidx["retry_policy"]] = "NONE"
            if "retry_max_count" in nidx:
                new_row[nidx["retry_max_count"]] = 0
            if "timeout_seconds" in nidx:
                new_row[nidx["timeout_seconds"]] = 0
            if "node_params" in nidx:
                new_row[nidx["node_params"]] = "{}"
            if "enabled" in nidx:
                new_row[nidx["enabled"]] = "TRUE"
            node_ws.append(new_row)
            appended += 1
            appended_codes.append("START")
        if not has_end:
            new_row = [None] * len(n_headers)
            new_row[nidx["tenant_id"]] = tenant
            new_row[nidx["workflow_code"]] = wf_code
            new_row[nidx["workflow_version"]] = wf_ver
            new_row[nidx["node_code"]] = "END"
            new_row[nidx["node_name"]] = "结束"
            new_row[nidx["node_type"]] = "END"
            new_row[nidx["node_order"]] = max_order + 1
            if "retry_policy" in nidx:
                new_row[nidx["retry_policy"]] = "NONE"
            if "retry_max_count" in nidx:
                new_row[nidx["retry_max_count"]] = 0
            if "timeout_seconds" in nidx:
                new_row[nidx["timeout_seconds"]] = 0
            if "node_params" in nidx:
                new_row[nidx["node_params"]] = "{}"
            if "enabled" in nidx:
                new_row[nidx["enabled"]] = "TRUE"
            node_ws.append(new_row)
            appended += 1
            appended_codes.append("END")

        # patch edges -> rename NODE_START/NODE_END and add START->first_job + last_job->END if missing
        if edge_ws is not None:
            e_headers = [c.value for c in edge_ws[1]]
            eidx = {h: i for i, h in enumerate(e_headers)}
            existing = set()
            for er in edge_ws.iter_rows(min_row=2):
                if er[eidx["tenant_id"]].value != tenant:
                    continue
                if er[eidx["workflow_code"]].value != wf_code:
                    continue
                if er[eidx["workflow_version"]].value != wf_ver:
                    continue
                # rename legacy boundary codes in edges
                if er[eidx["from_node_code"]].value == "NODE_START":
                    er[eidx["from_node_code"]].value = "START"
                if er[eidx["to_node_code"]].value == "NODE_END":
                    er[eidx["to_node_code"]].value = "END"
                existing.add((er[eidx["from_node_code"]].value,
                              er[eidx["to_node_code"]].value))

            def _add_edge(from_code, to_code):
                nonlocal edges_added
                if (from_code, to_code) in existing:
                    return
                er = [None] * len(e_headers)
                er[eidx["tenant_id"]] = tenant
                er[eidx["workflow_code"]] = wf_code
                er[eidx["workflow_version"]] = wf_ver
                er[eidx["from_node_code"]] = from_code
                er[eidx["to_node_code"]] = to_code
                if "edge_type" in eidx:
                    er[eidx["edge_type"]] = "SUCCESS"
                if "enabled" in eidx:
                    er[eidx["enabled"]] = "TRUE"
                edge_ws.append(er)
                existing.add((from_code, to_code))
                edges_added += 1

            if "START" in appended_codes and first_job:
                _add_edge("START", first_job)
            if "END" in appended_codes and last_job:
                _add_edge(last_job, "END")

    return renamed + appended, edges_added


def process(path: Path) -> dict:
    before = path.stat().st_size
    wb = load_workbook(path)
    result = {"file": path.name, "before_bytes": before}
    if "job_definition" in wb.sheetnames:
        ws = wb["job_definition"]
        result["schedule_expr_fixed"] = patch_schedule_expr(ws)
        result["execution_mode_fixed"] = ensure_execution_mode(ws)
    if "file_channel_config" in wb.sheetnames:
        result["sftp_keys_renamed"] = patch_channel_config_json(wb["file_channel_config"])
    if "file_template_config" in wb.sheetnames:
        result["file_template_config_fixed"] = patch_file_template_config(wb["file_template_config"])
    nodes_touched, edges_added = ensure_workflow_boundary_nodes(wb)
    result["workflow_nodes_touched"] = nodes_touched
    result["workflow_edges_added"] = edges_added
    wb.save(path)
    after = path.stat().st_size
    result["after_bytes"] = after
    return result


def verify(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    out = {"file": path.name}
    if "job_definition" in wb.sheetnames:
        ws = wb["job_definition"]
        headers = [c.value for c in ws[1]]
        col = headers.index("schedule_expr")
        type_col = headers.index("schedule_type") if "schedule_type" in headers else -1
        em_col = headers.index("execution_mode") if "execution_mode" in headers else -1
        bad = []
        missing_execution_mode = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            v = row[col]
            schedule_type = str(row[type_col] or "").strip().upper() if type_col >= 0 else "CRON"
            if schedule_type != "CRON" or v is None or str(v).strip() == "":
                pass
            else:
                n = len(str(v).split())
                if n not in (6, 7):
                    bad.append(v)
            if em_col < 0 or row[em_col] in (None, ""):
                missing_execution_mode += 1
        out["non_quartz_schedule_count"] = len(bad)
        out["missing_execution_mode_count"] = missing_execution_mode
    if "file_channel_config" in wb.sheetnames:
        ws = wb["file_channel_config"]
        headers = [c.value for c in ws[1]]
        ct = headers.index("channel_type"); cj = headers.index("config_json")
        bad = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if (row[ct] or "").upper() != "SFTP":
                continue
            try:
                obj = json.loads(row[cj])
            except Exception:
                continue
            if any(k in obj for k in ("host", "port", "username", "password", "remotePath")):
                bad += 1
        out["sftp_legacy_keys_left"] = bad
    if "file_template_config" in wb.sheetnames:
        ws = wb["file_template_config"]
        headers = [c.value for c in ws[1]]
        fm = headers.index("field_mappings")
        sql = headers.index("default_query_sql")
        missing_name = 0
        select_star = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw = row[fm]
            if raw not in (None, ""):
                try:
                    mappings = json.loads(str(raw))
                except Exception:
                    mappings = []
                if isinstance(mappings, list):
                    for mapping in mappings:
                        if isinstance(mapping, dict) and not str(mapping.get("name") or "").strip():
                            missing_name += 1
            query = str(row[sql] or "")
            if re.search(r"(?is)\bselect\s+\*", query):
                select_star += 1
        out["field_mapping_missing_name_count"] = missing_name
        out["export_select_star_count"] = select_star
    if "workflow_node" in wb.sheetnames:
        ws = wb["workflow_node"]
        headers = [c.value for c in ws[1]]
        nc = headers.index("node_code"); nt = headers.index("node_type")
        wfc = headers.index("workflow_code"); wfv = headers.index("workflow_version")
        seen = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[wfc]:
                continue
            k = (row[wfc], row[wfv])
            seen.setdefault(k, {"start": 0, "end": 0})
            if row[nt] == "START" and row[nc] == "START":
                seen[k]["start"] += 1
            if row[nt] == "END" and row[nc] == "END":
                seen[k]["end"] += 1
        out["workflow_boundary_ok"] = all(v["start"] == 1 and v["end"] == 1 for v in seen.values())
        out["workflows_checked"] = len(seen)
    return out


if __name__ == "__main__":
    for name in TARGETS:
        p = SUITE_DIR / name
        if not p.exists():
            print(f"MISSING {p}")
            continue
        r = process(p)
        print("PATCH", r)
        v = verify(p)
        print("VERIFY", v)
